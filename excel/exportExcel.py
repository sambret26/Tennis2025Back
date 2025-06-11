from datetime import datetime, timedelta
from openpyxl.styles import Font
import openpyxl
import io


from repositories.CompetitionRepository import CompetitionRepository
from repositories.CategoryRepository import CategoryRepository
from repositories.PlayerRepository import PlayerRepository
from repositories.CourtRepository import CourtRepository
from repositories.MatchRepository import MatchRepository
from repositories.TeamRepository import TeamRepository
from repositories.GridRepository import GridRepository
from constants import constants
from excel import utils

competitionRepository = CompetitionRepository()
categoryRepository = CategoryRepository()
playerRepository = PlayerRepository()
matchRepository = MatchRepository()
courtRepository = CourtRepository()
teamRepository = TeamRepository()
gridRepository = GridRepository()

def writeCell(sheet, column, row, value, fill=None, font=None):
    if value in ['0', 0, '', None]:
        value = ''
    sheet.cell(column=column, row=row, value=value)
    if fill:
        sheet.cell(column=column, row=row).fill = fill
    if font:
        sheet.cell(column=column, row=row).font = font

def addMatch(sheet, match, rowNum, gridsMap):
    player1 = getPlayer(match.label, rowNum, match.player1, match.futurPlayer1, match.double)
    player2 = getPlayer(match.label, rowNum, match.player2, match.futurPlayer2, match.double)
    writeCell(sheet, 1, rowNum, match.label)
    writeCell(sheet, 2, rowNum, player1)
    writeCell(sheet, 4, rowNum, player2)
    writeCell(sheet, 6, rowNum, utils.orZero(gridsMap.get(match.gridId)))
    if(match.player1Id and match.winnerId == match.player1Id):
        writeCell(sheet, 7, rowNum, player1)
    elif(match.player2Id and match.winnerId == match.player2Id):
        writeCell(sheet, 7, rowNum, player2)
    else:
        writeCell(sheet, 7, rowNum, 0)
    writeCell(sheet, 8, rowNum, utils.orZero(match.score))
    writeCell(sheet, 9, rowNum, utils.orZero(match.nextRound))

def getPlayer(matchLabel, rowNum, player, futurPlayer, double):
    if not player:
        if futurPlayer:
            if futurPlayer == "QE":
                return futurPlayer
            return "V" + futurPlayer
        return None
    #TODO : VS/VD
    #TODO : VT/VP/QE
    if double:
        return None #TODO : Double
    else:
        return player.getFullName()

def addFormulas(sheet):
    for rowNum in range (3, 200):
        writeCell(sheet, 3, rowNum, constants.MATCHS_FORMULA_COLUMN_C.replace("ROWNUM", str(rowNum)))
        writeCell(sheet, 5, rowNum, constants.MATCHS_FORMULA_COLUMN_E.replace("ROWNUM", str(rowNum)))

def getMatchesInOrder():
    allMatches = []
    for categorie in categoryRepository.getAllCategories():
        matches = matchRepository.getMatchesByCategorie(categorie.id)
        matches.sort(key=lambda m: int(m.label[2:]))
        for match in matches:
            allMatches.append(match)
    return allMatches

def addSlots(timesSlots, date, courtId):
    hours = [hour for hour, name, court in timesSlots]
    href = utils.getHref(date)
    hours.insert(0, href)
    max = 10
    for index, hour in enumerate(hours):
        if max == 0 :
            return timesSlots
        max -= 1
        if len(hours) > index+1:
            nextHour = hours[index+1]
        else:
            nextHour = constants.MAX_HOUR
        while utils.findDiff(hour, nextHour) > 179:
            nextHour = utils.plus90(hour)
            if utils.inMinutes(nextHour) > utils.inMinutes(href) + 89:
                hours.insert(index+1, nextHour)
                timesSlots.insert(index, (nextHour, 0, courtId))
            else:
                hours.insert(index+1, href)
                break
    return timesSlots

def addSchedule(sheet, day, courtsId, courtNameMap):
    matchesByCourts = []
    for courtId in courtsId:
        matches = matchRepository.getMatchesByDayAndCourt(day, courtId) #TODO Perf
        matchesByCourts.append((courtId, matches))
    timesSlots = []
    if utils.isNotPast(day):
        for (courtId, matchesInCourt) in matchesByCourts:
            courtName = courtNameMap[courtId]
            if courtName == "Court n°3":
                continue
            slots = addSlots(matchesInCourt, day, courtId)
            timesSlots.extend(slots)
    else:
        for (courtId, matchesInCourt) in matchesByCourts:
            for match in matchesInCourt:
                slots = (match.hour, match.label, courtId)
                timesSlots.append(slots)
    if(len(timesSlots) == 0):
        return
    timesSlots.sort(key=lambda ts: utils.inMinutes(ts[0]))
    firstRow = sheet.max_row + 1
    previousHour = None
    toMerge = False
    for rowNum, timeSlot in enumerate(timesSlots, start=firstRow):
        hour, matchLabel, courtId = timeSlot
        courtName = courtNameMap[courtId]
        rowNum = sheet.max_row + 1
        writeCell(sheet, 1, rowNum, utils.changeDate(day))
        writeCell(sheet, 2, rowNum, hour)
        writeCell(sheet, 3, rowNum, courtName)
        writeCell(sheet, 4, rowNum, matchLabel)
        for column in range(5, 13):
            writeCell(sheet, column, rowNum, constants.SCHEDULE_FORMULA.replace("ROWNUM", str(rowNum)).replace("COLUMN", str(column-3)))        
        if hour == previousHour and not toMerge:
            toMerge = True
            start_row = rowNum - 1
        elif hour != previousHour and toMerge:
            sheet.merge_cells(start_row=start_row, start_column=2, end_row=rowNum-1, end_column=2)
            toMerge = False
        previousHour = hour
    if toMerge:
        sheet.merge_cells(start_row=start_row, start_column=2, end_row=rowNum, end_column=2)
    sheet.merge_cells(start_row=firstRow, start_column=1, end_row=firstRow + len(timesSlots) - 1, end_column=1)

def initiatePlayersSheet(sheet):
    for colNum, header in enumerate(constants.PLAYERS_HEADERS, start=1):
        writeCell(sheet, colNum, 1, header)
    for colNum, width in enumerate(constants.PLAYERS_COLUMN_WIDTHS, start=1):
        sheet.column_dimensions[openpyxl.utils.get_column_letter(colNum)].width = width
    players = playerRepository.getAllPlayers()
    players.sort(key=lambda p: (p.rankingId, p.lastName))
    for rowNum, player in enumerate(players, start=2):
        writeCell(sheet, 1, rowNum, player.getFullName())
        writeCell(sheet, 2, rowNum, utils.orZero(player.ranking.simple if player.ranking else None))
    teams = teamRepository.getAllTeams()
    for rowNum, team in enumerate(teams, start=2):
        writeCell(sheet, 4, rowNum, team.getFullName())
        writeCell(sheet, 5, rowNum, team.ranking)
        writeCell(sheet, 6, rowNum, team.player1.firstName, None, Font(color="ffffff"))
        writeCell(sheet, 7, rowNum, team.player2.firstName, None, Font(color="ffffff"))

def initiateMatchesSheet(sheet):
    gridsMap = gridRepository.getGridsMap()
    for colNum, header in enumerate(constants.MATCHES_HEADERS, start=1):
        writeCell(sheet, colNum, 1, header)
    for colNum, width in enumerate(constants.MATCHES_COLUMN_WIDTHS, start=1):
        sheet.column_dimensions[openpyxl.utils.get_column_letter(colNum)].width = width
    matches = getMatchesInOrder()
    for rowNum, match in enumerate(matches, start=3):
        addMatch(sheet, match, rowNum, gridsMap)
    addFormulas(sheet)

def initiateScheduleSheet(sheet):
    for colNum, header in enumerate(constants.SCHEDULE_HEADERS, start=1):
        writeCell(sheet, colNum, 1, header)
    for colNum, width in enumerate(constants.SCHEDULE_COLUMN_WIDTHS, start=1):
        sheet.column_dimensions[openpyxl.utils.get_column_letter(colNum)].width = width
    startDay, endDay = competitionRepository.getDates()
    startDate = datetime.strptime(startDay, "%Y-%m-%d")
    endDate = datetime.strptime(endDay, "%Y-%m-%d")
    courtNameMap = courtRepository.getCourtNameMap()
    courtsId = courtRepository.getAllCourtId()
    for day in range((endDate - startDate).days + 1):
        currentDate = startDate + timedelta(days=day)
        addSchedule(sheet, currentDate.strftime("%Y-%m-%d"), courtsId, courtNameMap)

def createExcel():
    file = openpyxl.Workbook()
    defaultSheet = file['Sheet']
    file.remove(defaultSheet)
    matchesSheet = file.create_sheet(constants.MATCHES_SHEET_NAME)
    scheduleSheet = file.create_sheet(constants.SCHEDULE_SHEET_NAME)
    playersSheet = file.create_sheet(constants.PLAYERS_SHEET_NAME)
    initiatePlayersSheet(playersSheet)
    initiateMatchesSheet(matchesSheet)
    initiateScheduleSheet(scheduleSheet)
    utils.addCourtColor(scheduleSheet)
    utils.centerAndBorderSheetAndWhiteFontMatchs(matchesSheet)
    utils.centerAndBorderSheetAndWhiteFont(scheduleSheet)
    utils.centerAndBorderSheetAndWhiteFont(playersSheet, 5)
    excelBytes = io.BytesIO()
    file.save(excelBytes)
    excelBytes.seek(0)
    return excelBytes