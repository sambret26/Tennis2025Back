import locale
from datetime import datetime
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from constants import constants

def inMinutes(hour):
    if "H" in hour.upper():
        split = hour.upper().split("H")
    elif ":" in hour:
        split = hour.split(":")
    if len(split) == 1 :
        h = split[0]
    else :
        h,m = split
    h = int(h)
    if m == '': m = 0
    else: m = int(m)
    return h*60+m

def changeDate(date):
    try:
        locale.setlocale(locale.LC_TIME, 'C.UTF-8')
        parsed_date = datetime.strptime(date, '%Y-%m-%d')
        formatted_date = parsed_date.strftime('%A %d %B')
        return formatted_date.title()
    except ValueError:
        return date

def isNotPast(date):
    split = date.split("-")
    return int(datetime.now().strftime("%m%d")) < 100 * int(split[1]) + int(split[2])

def addCourtColor(sheet):
    rule1 = FormulaRule(formula=['AND(OR($C2=2,$C2="Court n°2"),D2<>0)'], fill=PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"))
    rule2 = FormulaRule(formula=['AND(OR($C2=2,$C2="Court n°2"),D2=0)'], fill=PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"), font=Font(color="D9D9D9"))
    rule3 = FormulaRule(formula=['AND(OR($C2=3,$C2="Court n°3"),D2<>0)'], fill=PatternFill(start_color="A6A6A6", end_color="A6A6A6", fill_type="solid"))
    rule4 = FormulaRule(formula=['AND(OR($C2=3,$C2="Court n°3"),D2=0)'], fill=PatternFill(start_color="A6A6A6", end_color="A6A6A6", fill_type="solid"), font=Font(color="A6A6A6"))
    for rule in [rule1, rule2, rule3, rule4]:
        sheet.conditional_formatting.add('D2:L1048576', rule)

def centerAndBorderSheetAndWhiteFont(sheet, maxColumn = None):
    borderType = Side(style='thin')
    border = Border(top=borderType, bottom=borderType, left=borderType, right=borderType)
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column if maxColumn == None else maxColumn):
        for cell in row:
            cell.alignment = Alignment(vertical="center", horizontal="center")
            if cell.value != None or cell.coordinate in sheet.merged_cells : cell.border = border
    addWhiteFont(sheet)

def centerAndBorderSheetAndWhiteFontMatchs(sheet):
    borderType = Side(style='thin')
    border = Border(top=borderType, bottom=borderType, left=borderType, right=borderType)
    row = 1
    while True and row < 200:
        if row == 2  :
            row = 3
            continue
        if sheet.cell(column = 1, row = row).value == None : break
        for col in range(1, 10):
            sheet.cell(column = col, row = row).alignment = Alignment(vertical="center", horizontal="center")
            sheet.cell(column = col, row = row).border = border
        row += 1
    addWhiteFont(sheet)

def addWhiteFont(sheet):
    rule = CellIsRule(operator='equal', formula=['0'], font=Font(color="ffffff"))
    sheet.conditional_formatting.add('A1:ZZ1048576', rule)

def getHref(date):
    date = datetime.strptime(date, "%Y-%m-%d")
    if date.weekday() >= 5:
        return constants.WEEKEND_START_HOUR
    return constants.WEEK_START_HOUR

def findDiff(hour, nextHour):
    return inMinutes(nextHour) - inMinutes(hour)

def plus90(hour):
    minutes = inMinutes(hour)
    minutes = minutes+90
    h = minutes//60
    m = (minutes - 60*h)
    if m ==0 : return f"{h}H"
    return f"{h}H{m}"

def orZero(value):
    return value if value != None else 0

def orNone(value):
    return value if value != 0 else None

def areDifferents(match1, match2):
    if match1['player1Id'] != match2.player1Id:
        return True
    if match1['player2Id'] != match2.player2Id:
        return True
    if match1['winnerId'] != match2.winnerId:
        return True
    if match1['panel'] != match2.panel:
        return True
    if match1['score'] != match2.score:
        return True
    if match1['nextRound'] != match2.nextRound:
        return True
    return False