import openpyxl
import io

from repositories.PlayerRepository import PlayerRepository
from repositories.MatchRepository import MatchRepository
from constants import constants
from excel import utils

playerRepository = PlayerRepository()
matchRepository = MatchRepository()

async def readMessage(message):
    data = await message.attachments[0].read()
    fileData = io.BytesIO(data)
    inputFile = openpyxl.load_workbook(fileData)
    if not await checkExcel(message.channel, inputFile):
        return
    handleExcel(inputFile)
    await message.channel.send("OK") #TODO Mettre en constante/ Traiter KO

async def checkExcel(channel, file):
    for sheetName in [constants.MATCHES_SHEET_NAME, constants.SCHEDULE_SHEET_NAME]:
        if not sheetName in file.sheetnames:
            await channel.send(constants.SHEET_NOT_FOUND.replace("SHEET_NAME", constants.sheetName))
            return False
    return True

def handleExcel(file):
    matchesToAdd = []
    matchesToUpdates = []
    matchesNamesToRemove = []
    errors = []
    handleMatchsSheet(file[constants.MATCHES_SHEET_NAME], matchesToAdd, matchesToUpdates, matchesNamesToRemove)
    handleScheduleSheet(file[constants.SCHEDULE_SHEET_NAME], matchesToAdd, matchesToUpdates, matchesNamesToRemove, errors)
    updateDB(matchesToAdd, matchesToUpdates, matchesNamesToRemove, errors)
    print("To add : \n")
    for match in matchesToAdd:
        print(str(match) + "\n")
    print("To update : \n")
    for match in matchesToUpdates:
        print(str(match) + "\n")

def handleMatchsSheet(sheet, toAdd, toUpdate, toRemove):
    row = 3
    listHandled = []
    matchsMap = matchRepository.getMatchesLabelMap()
    playersMap = playerRepository.getPlayersNamesMap()
    while True:
        matchName = readCell(sheet, 1, row)
        if matchName is None:
            break
        matchData = readData(sheet, row)
        matchData['player1Id'] = getPlayerIdByNameAndMatchType(playersMap, matchData["player1"], matchName)
        matchData['player2Id'] = getPlayerIdByNameAndMatchType(playersMap, matchData["player2"], matchName)
        matchData['winnerId'] = getWinnerId(matchData)
        matchInfosInDB = matchsMap.get(matchName)
        if matchsInfosInDB is None:
            toAdd.append(matchData) #TODO Adapter ce qu'on met dans la liste
            continue
        listHandled.append(f'{matchName}')
        if utils.areDifferents(matchData, matchsInfosInDB):
            toUpdate.append(matchData) #TODO Adapter ce qu'on met dans la liste
        row += 1
    toRemove = matchRepository.getMatchesNamesToDelete(listHandled)

def handleScheduleSheet():
    return #TODO

def updateDB():
    return #TODO

def getPlayerIdByNameAndMatchType(playersMap, player, matchName):
    if player in ["", " ", None]:
        return None
    if player.startswith("VS") or player.startswith("VD") or player.startswith("VT") or player.startswith("VT") or player.startswith("QE"): #TODO : Remove VT/VP ?
        return player
    if player.startswith("=IF"):
        if ('\"') in player:
            return player.split('\"')[1]
        return None
    if matchName.startswith("S"):
        split = player.split(' ')
        playerName = ' '.join(split[0:-1])
        playerFirstName = split[-1]
        fullName = playerName + '_' + playerFirstName
        return str(playersMap.get(fullName))
    return None #TODO Doubles

def getWinnerId(matchData):
    if matchData['winner'] == matchData['player1']:
        return matchData['player1Id']
    if matchData['winner'] == matchData['player2']:
        return matchData['player2Id']
    return None

def readData(sheet, row):
    return {
        "player1": readCell(sheet, 2, row),
        "player2": readCell(sheet, 4, row),
        "panel": utils.orNone(readCell(sheet, 6, row)),
        "winner": readCell(sheet, 7, row),
        "score": utils.orNone(readCell(sheet, 8, row)),
        "nextRound": utils.orNone(readCell(sheet, 9, row))
    }

def readCell(sheet, column, row):
    return sheet.cell(column=column, row=row).value